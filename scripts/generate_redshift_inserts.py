#!/usr/bin/env python3
"""Generate Redshift INSERTs from LCC Excel + table_data_mapping.xlsx."""

from __future__ import annotations

import math
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA_XLSX = ROOT / "LCC - Inventory Liquidation Data.xlsx"
MAPPING_XLSX = ROOT / "table_data_mapping.xlsx"
OUT_SQL = ROOT / "generated_inserts.sql"
OUT_REPORT = ROOT / "mapping_coverage_report.txt"

# Full column order per inventory.* DDL (schema.sql)
TABLE_COLUMNS: dict[str, list[str]] = {
    "inventory.warehouses": [
        "id",
        "name",
        "location_code",
        "city",
        "region",
        "pincode",
        "stock_units",
        "capacity_pct",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.master_sink_config": [
        "id",
        "warehouse_id",
        "warehouse_name",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.donor_settings": [
        "warehouse_id",
        "is_participating",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.route_pair_overrides": [
        "donor_warehouse_id",
        "sink_warehouse_id",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.categories": [
        "id",
        "name",
        "shelf_life_override_pct",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.brands": [
        "id",
        "name",
        "category_id",
        "shelf_life_override_pct",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.skus": [
        "id",
        "name",
        "brand_id",
        "brand_name",
        "category_id",
        "category_name",
        "type",
        "shelf_life_override_pct",
        "is_ignored",
        "is_active",
        "stock_units",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.thresholds_global": [
        "id",
        "cogs_min",
        "cogs_max",
        "units_min",
        "units_max",
        "weight_min",
        "weight_max",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.thresholds_category": [
        "category_id",
        "category_name",
        "cogs_min",
        "cogs_max",
        "units_min",
        "units_max",
        "weight_min",
        "weight_max",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.thresholds_brand": [
        "brand_id",
        "brand_name",
        "category_id",
        "category_name",
        "cogs_min",
        "cogs_max",
        "units_min",
        "units_max",
        "weight_min",
        "weight_max",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.product_config_global": [
        "id",
        "standard_shelf_life_pct",
        "op_shelf_life_pct",
        "standard_enabled",
        "op_enabled",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.inventory_conditions": [
        "id",
        "condition_type",
        "description",
        "is_enabled",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.batch_runs": [
        "id",
        "master_sink_id",
        "master_sink_name",
        "status",
        "generated_at",
        "committed_by",
        "committed_at",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
    "inventory.batches": [
        "id",
        "master_id",
        "wh_id",
        "wh_name",
        "status",
        "generated_at",
        "committed_by",
        "committed_at",
        "is_active",
        "created_by",
        "create_dt",
        "updated_by",
        "update_dt",
    ],
}

BOOL_COLS = {
    "is_active",
    "is_participating",
    "is_ignored",
    "is_enabled",
    "standard_enabled",
    "op_enabled",
}

INT_COLS = {
    "stock_units",
    "units_min",
    "units_max",
}

DECIMAL_COLS = {
    "cogs_min",
    "cogs_max",
    "weight_min",
    "weight_max",
    "capacity_pct",
    "shelf_life_override_pct",
    "standard_shelf_life_pct",
    "op_shelf_life_pct",
}

TS_COLS = {
    "create_dt",
    "update_dt",
    "generated_at",
    "committed_at",
}


def norm_key(h: str | None) -> str:
    if h is None:
        return ""
    return str(h).strip().lower().replace(" ", "_")


def escape_sql_str(s: str) -> str:
    return s.replace("'", "''")


def excel_headers_to_map(header_row: tuple) -> dict[str, int]:
    m: dict[str, int] = {}
    for i, h in enumerate(header_row):
        k = norm_key(h if h is not None else None)
        if k:
            m[k] = i
    return m


def coerce_bool(v) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return "TRUE" if v != 0 else "FALSE"
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "t"):
        return "TRUE"
    if s in ("false", "0", "no", "f", ""):
        return "FALSE"
    return "NULL"


def format_sql_value(col: str, v, table: str) -> str:
    if v is None or (isinstance(v, str) and v.strip() == "" and col in ("updated_by", "update_dt", "description")):
        if col in ("updated_by", "update_dt", "committed_by", "committed_at", "description"):
            return "NULL"
    if v is None:
        return "NULL"
    if col in BOOL_COLS:
        return coerce_bool(v)
    if col in INT_COLS:
        if isinstance(v, float) and v == int(v):
            v = int(v)
        try:
            if v == "" or v is None:
                return "NULL"
            return str(int(float(v)))
        except (TypeError, ValueError):
            return "NULL"
    if col in DECIMAL_COLS:
        try:
            if v == "" or v is None:
                return "NULL"
            return str(float(v))
        except (TypeError, ValueError):
            return "NULL"
    if col in TS_COLS or col.endswith("_at"):
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return "NULL"
        if hasattr(v, "isoformat"):
            s = v.isoformat()
        else:
            s = str(v).strip()
        if "T" in s:
            s = s.replace("T", " ")
        if s.endswith("Z"):
            s = s[:-1]
        # strip fractional seconds if problematic (keep ms)
        return f"TIMESTAMP '{escape_sql_str(s)}'"
    # varchar-ish id columns: strip float .0
    if col in ("id", "warehouse_id", "donor_warehouse_id", "sink_warehouse_id", "brand_id", "category_id", "master_id", "wh_id", "master_sink_id"):
        if isinstance(v, float) and v == int(v):
            v = int(v)
        s = str(v).strip()
        return f"'{escape_sql_str(s)}'"
    if col in ("pincode", "location_code"):
        if isinstance(v, float) and v == int(v):
            v = int(v)
        s = str(v).strip()
        return f"'{escape_sql_str(s)}'"
    if isinstance(v, float) and v == int(v) and col == "id" and "warehouse" in table:
        v = int(v)
    if isinstance(v, (int, float)) and col not in ("name",):
        if isinstance(v, float) and abs(v - round(v)) < 1e-9 and col in ("id", "pincode"):
            return f"'{escape_sql_str(str(int(round(v))))}'"
    s = str(v).strip() if v is not None else ""
    return f"'{escape_sql_str(s)}'"


def resolve_sheet_name(requested: str | None, data_sheetnames: list[str]) -> str | None:
    if not requested:
        return None
    r = requested.strip()
    # typo in mapping file
    if r == "InventoryConditons":
        r = "InventoryConditions"
    lower_map = {s.lower(): s for s in data_sheetnames}
    return lower_map.get(r.lower())


def read_mapping() -> list[tuple[str, str | None]]:
    wb = openpyxl.load_workbook(MAPPING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out: list[tuple[str, str | None]] = []
    for i, row in enumerate(rows):
        if i == 0:
            continue
        t, s = row[0], row[1] if len(row) > 1 else None
        if t:
            out.append((str(t).strip(), str(s).strip() if s else None))
    return out


def sheet_rows(sheet_name: str) -> tuple[list[str], list[tuple]]:
    wb = openpyxl.load_workbook(DATA_XLSX, read_only=True, data_only=True)
    ws = wb[sheet_name]
    it = ws.iter_rows(values_only=True)
    header = next(it)
    cols = [str(c).strip() if c is not None else "" for c in header]
    data = [tuple(r) for r in it]
    wb.close()
    return cols, data


def row_is_empty(row: tuple, col_idx: dict[str, int], min_keys: list[str]) -> bool:
    for k in min_keys:
        i = col_idx.get(norm_key(k))
        if i is None:
            continue
        v = row[i] if i < len(row) else None
        if v is not None and str(v).strip() != "":
            return False
    return True


def cell_non_empty(v) -> bool:
    if v is None:
        return False
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return False
    if isinstance(v, str) and not v.strip():
        return False
    return True


def required_fields_for_skip(table: str) -> list[str]:
    """Excel columns that must be non-blank; avoids INSERT NULL into NOT NULL columns."""
    if table == "inventory.donor_settings":
        return ["warehouse_id", "created_by"]
    if table == "inventory.route_pair_overrides":
        return ["donor_warehouse_id", "sink_warehouse_id", "created_by"]
    if table == "inventory.skus":
        return ["id", "name", "brand_id", "category_id", "type", "created_by"]
    cols = TABLE_COLUMNS.get(table, [])
    if "id" in cols:
        return ["id"]
    return []


def row_should_skip(table: str, col_idx: dict[str, int], row: tuple) -> bool:
    for f in required_fields_for_skip(table):
        j = col_idx.get(norm_key(f))
        if j is None:
            continue
        v = row[j] if j < len(row) else None
        if not cell_non_empty(v):
            return True
    return False


def batches_transform(header: list[str], row: tuple) -> tuple[list[str], tuple]:
    """Map Excel 'Batches' sheet (batch_runs-shaped) onto inventory.batches DDL."""
    idx = {norm_key(h): j for j, h in enumerate(header)}
    # Build synthetic row dict for batches columns
    def get(*names):
        for n in names:
            j = idx.get(norm_key(n))
            if j is not None and j < len(row):
                return row[j]
        return None

    master_sink_id = get("master_sink_id")
    master_sink_name = get("master_sink_name")
    rid = get("id")
    # Heuristic: Excel row matches batch_runs layout; batches needs donor wh + run id.
    vals = {
        "id": rid,
        "master_id": master_sink_id,  # NOTE: wrong FK semantics — see report
        "wh_id": master_sink_id,
        "wh_name": master_sink_name,
        "status": get("status"),
        "generated_at": get("generated_at"),
        "committed_by": get("committed_by"),
        "committed_at": get("committed_at"),
        "is_active": get("is_active"),
        "created_by": get("created_by"),
        "create_dt": get("create_dt"),
        "updated_by": get("updated_by"),
        "update_dt": get("update_dt"),
    }
    new_header = list(TABLE_COLUMNS["inventory.batches"])
    new_row = tuple(vals[c] for c in new_header)
    return new_header, new_row


def main() -> None:
    wb_probe = openpyxl.load_workbook(DATA_XLSX, read_only=True, data_only=True)
    data_sheets = list(wb_probe.sheetnames)
    wb_probe.close()

    mapping = read_mapping()
    mapped_tables = {t for t, _ in mapping}
    all_schema_tables = set(TABLE_COLUMNS)

    used_sheets: set[str] = set()
    sql_blocks: list[str] = []
    warnings: list[str] = []

    sql_blocks.append(
        "-- Generated by scripts/generate_redshift_inserts.py\n"
        "-- Target: Redshift schema `inventory`\n"
        "BEGIN;\n"
    )

    CHUNK = 200

    for table, sheet_ref in mapping:
        if table not in TABLE_COLUMNS:
            warnings.append(f"Unknown table in mapping (no DDL in script): {table}")
            continue
        sn = resolve_sheet_name(sheet_ref, data_sheets)
        if not sn:
            warnings.append(f"No sheet for {table} (mapping sheet ref: {sheet_ref!r})")
            continue
        used_sheets.add(sn)
        cols_ddl = TABLE_COLUMNS[table]

        header, data = sheet_rows(sn)
        col_idx = excel_headers_to_map(tuple(header))

        if table == "inventory.batches":
            warnings.append(
                "inventory.batches: Excel sheet 'Batches' columns match `inventory.batch_runs`, not `inventory.batches`. "
                "Generated SQL maps master_sink_id -> master_id and wh_id for load only; fix data or mapping before production."
            )

        rows_out: list[tuple] = []
        skipped_bad = 0
        for row in data:
            if table == "inventory.thresholds_category":
                if row_is_empty(row, col_idx, ["category_id", "category_name"]):
                    continue
            if table == "inventory.thresholds_brand":
                if row_is_empty(row, col_idx, ["brand_id"]):
                    continue
            if table == "inventory.batches":
                _, mapped_row = batches_transform(header, row)
                if not cell_non_empty(mapped_row[0]):
                    skipped_bad += 1
                    continue
                rows_out.append(mapped_row)
            else:
                if row_should_skip(table, col_idx, row):
                    skipped_bad += 1
                    continue
                rows_out.append(row)

        if skipped_bad:
            warnings.append(f"{table}: skipped {skipped_bad} sheet row(s) with blank required key(s)")

        if not rows_out:
            sql_blocks.append(f"\n-- {table} from sheet {sn!r}: no data rows\n")
            continue

        # INSERT in chunks
        for start in range(0, len(rows_out), CHUNK):
            chunk = rows_out[start : start + CHUNK]
            col_list = ", ".join(cols_ddl)
            value_rows: list[str] = []
            for row in chunk:
                if table != "inventory.batches":
                    values = []
                    for c in cols_ddl:
                        nk = norm_key(c)
                        j = col_idx.get(nk)
                        if j is None:
                            values.append("NULL")
                        else:
                            v = row[j] if j < len(row) else None
                            values.append(format_sql_value(c, v, table))
                    value_rows.append("(" + ", ".join(values) + ")")
                else:
                    values = []
                    for i, c in enumerate(cols_ddl):
                        values.append(format_sql_value(c, row[i], table))
                    value_rows.append("(" + ", ".join(values) + ")")

            sql_blocks.append(
                f"\n-- {table} <= sheet {sn!r} rows {start + 1}-{start + len(chunk)} of {len(rows_out)}\n"
                f"INSERT INTO {table} ({col_list})\nVALUES\n"
                + ",\n".join(value_rows)
                + ";\n"
            )

    sql_blocks.append("\nCOMMIT;\n")

    unused_sheets = sorted(set(data_sheets) - used_sheets)
    tables_in_mapping = {t for t, _ in mapping}
    tables_not_in_mapping_file = sorted(all_schema_tables - tables_in_mapping)
    tables_with_no_sheet = sorted({t for t, sref in mapping if not sref or not str(sref).strip()})

    report_lines = [
        "=== Sheets in LCC - Inventory Liquidation Data.xlsx not referenced by table_data_mapping ===",
        *(f"  - {s}" for s in unused_sheets),
        "",
        "=== inventory.* tables in DDL with no entry in table_data_mapping ===",
        *(f"  - {t}" for t in tables_not_in_mapping_file),
        "",
        "=== Tables listed in table_data_mapping with no sheet (NULL / empty sheet_reference) ===",
        *(f"  - {t}" for t in tables_with_no_sheet),
        "",
        "=== Warnings ===",
        *(f"  - {w}" for w in warnings),
        "",
        "=== Mapping typo note ===",
        "  - Mapping uses 'InventoryConditons'; resolved to sheet 'InventoryConditions'.",
    ]

    OUT_SQL.write_text("".join(sql_blocks), encoding="utf-8")
    OUT_REPORT.write_text("\n".join(report_lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_SQL}")
    print(f"Wrote {OUT_REPORT}")


if __name__ == "__main__":
    main()

